"""Apply formatting (font, color, borders, alignment) to Excel cells."""

import argparse
import json
import sys
import os

sys.path.insert(0, os.path.dirname(__file__))
from excel_utils import (
    get_app, get_workbook, get_sheet,
    open_file_writable, get_sheet_openpyxl,
    hex_to_rgb_int, output_json, IS_WINDOWS
)

# Excel alignment constants (for xlwings/COM)
H_ALIGN = {'left': -4131, 'center': -4108, 'right': -4152}
V_ALIGN = {'top': -4160, 'middle': -4108, 'bottom': -4107}

# Excel border constants (for xlwings/COM)
BORDER_POSITIONS = {
    'left': 7, 'right': 10, 'top': 8, 'bottom': 9,
    'inside_vertical': 11, 'inside_horizontal': 12
}
BORDER_STYLES = {
    'none': 0, 'thin': 1, 'medium': -4138, 'thick': 4,
    'double': -4119, 'dotted': -4118, 'dashed': -4115
}


def _parse_hex(hex_color):
    """Convert #RRGGBB to (r, g, b) tuple."""
    if hex_color.startswith('#'):
        hex_color = hex_color[1:]
    return (int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16))


# ---------------------------------------------------------------------------
# xlwings (live Excel)
# ---------------------------------------------------------------------------

def _format_live(workbook, cell_range, fmt, sheet):
    app, err = get_app()
    if err:
        return {"error": err}
    wb, err = get_workbook(app, workbook)
    if err:
        return {"error": err}
    ws, err = get_sheet(wb, sheet)
    if err:
        return {"error": err}

    try:
        rng = ws.range(cell_range)
    except Exception as e:
        return {"error": f"Invalid range '{cell_range}': {e}"}

    try:
        # Cross-platform xlwings API
        if 'bold' in fmt:
            rng.font.bold = fmt['bold']
        if 'italic' in fmt:
            rng.font.italic = fmt['italic']
        if 'fontSize' in fmt:
            rng.font.size = fmt['fontSize']
        if 'fontName' in fmt:
            rng.font.name = fmt['fontName']
        if 'fontColor' in fmt:
            rng.font.color = _parse_hex(fmt['fontColor'])
        if 'backgroundColor' in fmt:
            rng.color = _parse_hex(fmt['backgroundColor'])
        if 'numberFormat' in fmt:
            rng.number_format = fmt['numberFormat']

        # Platform-specific: underline
        if 'underline' in fmt:
            val = 2 if fmt['underline'] else -4142
            try:
                if IS_WINDOWS:
                    rng.api.Font.Underline = val
                else:
                    rng.api.font_object.underline.set(val)
            except Exception:
                pass

        # Platform-specific: alignment
        h = fmt.get('textAlign')
        v = fmt.get('verticalAlign')
        wrap = fmt.get('wrapText')
        if h and h in H_ALIGN:
            try:
                if IS_WINDOWS:
                    rng.api.HorizontalAlignment = H_ALIGN[h]
                else:
                    rng.api.horizontal_alignment.set(H_ALIGN[h])
            except Exception:
                pass
        if v and v in V_ALIGN:
            try:
                if IS_WINDOWS:
                    rng.api.VerticalAlignment = V_ALIGN[v]
                else:
                    rng.api.vertical_alignment.set(V_ALIGN[v])
            except Exception:
                pass
        if wrap is not None:
            try:
                if IS_WINDOWS:
                    rng.api.WrapText = wrap
                else:
                    rng.api.wrap_text.set(wrap)
            except Exception:
                pass

        # Borders
        if 'borders' in fmt:
            _borders_live(rng, fmt['borders'])

        try:
            app.screen_updating = True
        except Exception:
            pass

        return {"success": True, "workbook": wb.name, "sheet": ws.name, "range": cell_range}
    except Exception as e:
        return {"error": f"Failed to format: {e}"}


def _borders_live(rng, borders_config):
    for position, config in borders_config.items():
        if not config:
            continue
        style = config.get('style', 'thin')
        color = config.get('color', '#000000')
        rgb = hex_to_rgb_int(color)
        line_style = BORDER_STYLES.get(style, 1)

        if position == 'outside':
            positions = ['left', 'right', 'top', 'bottom']
        elif position == 'inside':
            positions = ['inside_vertical', 'inside_horizontal']
        elif position in BORDER_POSITIONS:
            positions = [position]
        else:
            continue

        for pos in positions:
            idx = BORDER_POSITIONS.get(pos)
            if idx is None:
                continue
            try:
                if IS_WINDOWS:
                    border = rng.api.Borders(idx)
                    border.LineStyle = line_style
                    if style != 'none':
                        border.Color = rgb
                else:
                    border = rng.api.borders[idx]
                    border.line_style.set(line_style)
                    if style != 'none':
                        border.color.set(rgb)
            except Exception:
                pass


# ---------------------------------------------------------------------------
# openpyxl (file-based)
# ---------------------------------------------------------------------------

def _format_file(path, cell_range, fmt, sheet):
    wb, err = open_file_writable(path)
    if err:
        return {"error": err}

    ws, err = get_sheet_openpyxl(wb, sheet)
    if err:
        wb.close()
        return {"error": err}

    try:
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except Exception as e:
        wb.close()
        return {"error": f"Invalid range '{cell_range}': {e}"}

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Build style objects
    font_kwargs = {}
    if 'bold' in fmt:
        font_kwargs['bold'] = fmt['bold']
    if 'italic' in fmt:
        font_kwargs['italic'] = fmt['italic']
    if 'underline' in fmt:
        font_kwargs['underline'] = 'single' if fmt['underline'] else 'none'
    if 'fontSize' in fmt:
        font_kwargs['size'] = fmt['fontSize']
    if 'fontName' in fmt:
        font_kwargs['name'] = fmt['fontName']
    if 'fontColor' in fmt:
        font_kwargs['color'] = fmt['fontColor'].lstrip('#')

    fill = None
    if 'backgroundColor' in fmt:
        fill = PatternFill(start_color=fmt['backgroundColor'].lstrip('#'),
                           end_color=fmt['backgroundColor'].lstrip('#'),
                           fill_type='solid')

    align_kwargs = {}
    if 'textAlign' in fmt:
        align_kwargs['horizontal'] = fmt['textAlign']
    if 'verticalAlign' in fmt:
        align_kwargs['vertical'] = fmt['verticalAlign']
    if 'wrapText' in fmt:
        align_kwargs['wrap_text'] = fmt['wrapText']

    nf = fmt.get('numberFormat')

    border = None
    if 'borders' in fmt:
        border = _build_openpyxl_border(fmt['borders'])

    try:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if font_kwargs:
                    # Merge with existing font to preserve unset properties
                    existing = cell.font
                    merged = {
                        'name': font_kwargs.get('name', existing.name),
                        'size': font_kwargs.get('size', existing.size),
                        'bold': font_kwargs.get('bold', existing.bold),
                        'italic': font_kwargs.get('italic', existing.italic),
                        'underline': font_kwargs.get('underline', existing.underline),
                        'color': font_kwargs.get('color', existing.color),
                    }
                    cell.font = Font(**merged)
                if fill:
                    cell.fill = fill
                if align_kwargs:
                    cell.alignment = Alignment(**align_kwargs)
                if nf:
                    cell.number_format = nf
                if border:
                    cell.border = border

        wb.save(path)
    except Exception as e:
        wb.close()
        return {"error": f"Failed to format: {e}"}

    wb.close()
    return {"success": True, "path": path, "sheet": ws.title, "range": cell_range}


def _build_openpyxl_border(borders_config):
    """Build an openpyxl Border object from our config format."""
    from openpyxl.styles import Border, Side

    sides = {}
    for position, config in borders_config.items():
        if not config:
            continue
        style = config.get('style', 'thin')
        color = config.get('color', '000000').lstrip('#')
        side = Side(style=style, color=color)

        if position == 'outside':
            sides.setdefault('left', side)
            sides.setdefault('right', side)
            sides.setdefault('top', side)
            sides.setdefault('bottom', side)
        elif position == 'inside':
            # openpyxl doesn't have inside borders per-cell;
            # apply to all edges as approximation
            sides.setdefault('left', side)
            sides.setdefault('right', side)
            sides.setdefault('top', side)
            sides.setdefault('bottom', side)
        else:
            sides[position] = side

    return Border(**sides)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--workbook', default=None)
    parser.add_argument('--path', default=None)
    parser.add_argument('--range', required=True)
    parser.add_argument('--format', required=True)
    parser.add_argument('--sheet', default=None)
    args = parser.parse_args()

    if not args.workbook and not args.path:
        output_json({"error": "Either --workbook or --path is required"})
        return

    try:
        fmt = json.loads(args.format)
    except json.JSONDecodeError:
        output_json({"error": "Invalid JSON for format"})
        return

    if args.path:
        result = _format_file(args.path, args.range, fmt, args.sheet)
    else:
        result = _format_live(args.workbook, args.range, fmt, args.sheet)

    output_json(result)


if __name__ == "__main__":
    main()
