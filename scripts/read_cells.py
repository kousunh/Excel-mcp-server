"""Read cell values and optionally formatting from an Excel range."""

import argparse
import sys
import os
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(__file__))
from excel_utils import (
    get_app, get_workbook, get_sheet, open_path,
    rgb_tuple_to_hex, output_json, IS_WINDOWS
)


def clean_value(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, float) and val == int(val):
        return int(val)
    if isinstance(val, str):
        return val.replace('\x00', '').strip()
    return val


# ---------------------------------------------------------------------------
# xlwings (live Excel / workbook mode)
# ---------------------------------------------------------------------------

def _read_live(workbook, cell_range, sheet, include_formats):
    app, err = get_app()
    if err:
        return {"error": err}
    wb, err = get_workbook(app, workbook)
    if err:
        return {"error": err}
    ws, err = get_sheet(wb, sheet)
    if err:
        return {"error": err}

    try:
        rng = ws.range(cell_range)
    except Exception as e:
        return {"error": f"Invalid range '{cell_range}': {e}"}

    values = _xlwings_values(rng)

    result = {"workbook": wb.name, "sheet": ws.name, "range": cell_range, "values": values}

    if include_formats:
        top_left = rng[0, 0]
        bottom_right = rng[-1, -1] if rng.shape[0] > 1 or rng.shape[1] > 1 else top_left
        result["formats"] = _read_formats_live(
            ws, top_left.row, top_left.column,
            bottom_right.row, bottom_right.column
        )
    return result


def _xlwings_values(rng):
    raw = rng.value
    if rng.shape[0] == 1 and rng.shape[1] == 1:
        return [[clean_value(raw)]]
    elif rng.shape[0] == 1:
        return [[clean_value(v) for v in raw]]
    elif rng.shape[1] == 1:
        return [[clean_value(v)] for v in raw]
    else:
        return [[clean_value(v) for v in row] for row in raw]


def _read_formats_live(sheet, r1, c1, r2, c2):
    formats = []
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            try:
                cell = sheet.range(row, col)
                fmt = {}
                bg = rgb_tuple_to_hex(cell.color)
                if bg:
                    fmt["bg"] = bg
                try:
                    if cell.font.bold:
                        fmt["bold"] = True
                    if cell.font.italic:
                        fmt["italic"] = True
                    if cell.font.size:
                        fmt["fontSize"] = cell.font.size
                    if cell.font.name:
                        fmt["fontName"] = cell.font.name
                    fc = rgb_tuple_to_hex(cell.font.color)
                    if fc and fc != "#000000":
                        fmt["fontColor"] = fc
                except Exception:
                    pass
                try:
                    nf = cell.number_format
                    if nf and nf != "General":
                        fmt["numberFormat"] = nf
                except Exception:
                    pass
                try:
                    borders = _borders_live(cell)
                    if borders:
                        fmt["borders"] = borders
                except Exception:
                    pass
                try:
                    align = _alignment_live(cell)
                    if align:
                        fmt.update(align)
                except Exception:
                    pass
                if fmt:
                    fmt["cell"] = cell.address.replace('$', '')
                    formats.append(fmt)
            except Exception:
                continue
    return formats


def _borders_live(cell):
    borders = {}
    names_indices = [("top", 8), ("bottom", 9), ("left", 7), ("right", 10)]
    style_map = {1: "thin", -4138: "medium", 4: "thick", -4119: "double", -4118: "dotted", -4115: "dashed"}
    for name, idx in names_indices:
        try:
            if IS_WINDOWS:
                ls = cell.api.Borders(idx).LineStyle
            else:
                ls = cell.api.borders[idx].line_style()
            if ls is not None and ls != -4142 and ls != 0:
                borders[name] = style_map.get(ls, "thin")
        except Exception:
            continue
    return borders or None


def _alignment_live(cell):
    result = {}
    h_map = {-4131: "left", -4108: "center", -4152: "right"}
    v_map = {-4160: "top", -4108: "middle", -4107: "bottom"}
    try:
        if IS_WINDOWS:
            h, v = cell.api.HorizontalAlignment, cell.api.VerticalAlignment
        else:
            h, v = cell.api.horizontal_alignment(), cell.api.vertical_alignment()
        if h in h_map:
            result["textAlign"] = h_map[h]
        if v in v_map and v_map[v] != "bottom":
            result["verticalAlign"] = v_map[v]
    except Exception:
        pass
    return result or None


# ---------------------------------------------------------------------------
# openpyxl (file-based read-only, no Excel needed)
# ---------------------------------------------------------------------------

def _read_file(path, cell_range, sheet, include_formats):
    import openpyxl
    from openpyxl.utils import range_boundaries

    if not os.path.exists(path):
        return {"error": f"File not found: {path}"}

    try:
        wb = openpyxl.load_workbook(path, read_only=not include_formats, data_only=True)
    except Exception as e:
        return {"error": f"Cannot open file: {e}"}

    # Get sheet
    if sheet:
        if sheet not in wb.sheetnames:
            wb.close()
            return {"error": f"Sheet '{sheet}' not found"}
        ws = wb[sheet]
    else:
        ws = wb.active

    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except Exception as e:
        wb.close()
        return {"error": f"Invalid range '{cell_range}': {e}"}

    values = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        values.append([clean_value(cell.value) for cell in row])

    result = {"path": path, "sheet": ws.title, "range": cell_range, "values": values}

    if include_formats:
        result["formats"] = _read_formats_file(ws, min_row, min_col, max_row, max_col)

    wb.close()
    return result


def _read_formats_file(ws, r1, c1, r2, c2):
    from openpyxl.utils import get_column_letter

    formats = []
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            fmt = {}

            # Background
            try:
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != '00000000':
                    rgb = str(fill.fgColor.rgb)
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    fmt["bg"] = f"#{rgb.lower()}"
            except Exception:
                pass

            # Font
            try:
                font = cell.font
                if font.bold:
                    fmt["bold"] = True
                if font.italic:
                    fmt["italic"] = True
                if font.size:
                    fmt["fontSize"] = font.size
                if font.name:
                    fmt["fontName"] = font.name
                if font.color and font.color.rgb:
                    rgb = str(font.color.rgb)
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    if rgb.lower() != "000000":
                        fmt["fontColor"] = f"#{rgb.lower()}"
            except Exception:
                pass

            # Number format
            try:
                nf = cell.number_format
                if nf and nf != "General":
                    fmt["numberFormat"] = nf
            except Exception:
                pass

            # Borders
            try:
                border = cell.border
                borders = {}
                for side_name in ("top", "bottom", "left", "right"):
                    side = getattr(border, side_name, None)
                    if side and side.style and side.style != "none":
                        borders[side_name] = side.style
                if borders:
                    fmt["borders"] = borders
            except Exception:
                pass

            # Alignment
            try:
                align = cell.alignment
                if align.horizontal and align.horizontal != "general":
                    fmt["textAlign"] = align.horizontal
                if align.vertical and align.vertical not in ("bottom", None):
                    fmt["verticalAlign"] = align.vertical
            except Exception:
                pass

            if fmt:
                fmt["cell"] = f"{get_column_letter(col)}{row}"
                formats.append(fmt)

    return formats


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--workbook', default=None)
    parser.add_argument('--path', default=None)
    parser.add_argument('--range', required=True)
    parser.add_argument('--sheet', default=None)
    parser.add_argument('--formats', action='store_true')
    args = parser.parse_args()

    if not args.workbook and not args.path:
        output_json({"error": "Either --workbook or --path is required"})
        return

    if args.path:
        result = _read_file(args.path, args.range, args.sheet, args.formats)
    else:
        result = _read_live(args.workbook, args.range, args.sheet, args.formats)

    output_json(result)


if __name__ == "__main__":
    main()
